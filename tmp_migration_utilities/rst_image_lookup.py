#!/usr/bin/env python3
"""
rst_image_lookup.py

Given a wiki.analog.com page URL, prints the ready-to-paste RST ``.. figure::``
directives for every content image on that page, with the correct relative path
from the migrated RST file location inside docs/.

Usage:
    python3 rst_image_lookup.py <wiki-page-url> [<wiki-page-url> ...]

    python3 rst_image_lookup.py \\
        https://wiki.analog.com/resources/eval/user-guides/ad-fmcomms1-ebz/hardware

    # Fuzzy / partial match (substring of URL):
    python3 rst_image_lookup.py ad-fmcomms1-ebz/hardware

    # List all pages that have content images (useful for exploring):
    python3 rst_image_lookup.py --list-pages

    # Filter listed pages by substring:
    python3 rst_image_lookup.py --list-pages eval/user-guides

Options:
    --list-pages [filter]   List all wiki pages that have downloadable images.
    --docs-root PATH        Root of the docs tree (default: docs/ next to this script).
    --no-figure             Emit ``.. image::`` instead of ``.. figure::``.
"""

import os
import sys
import argparse
from collections import defaultdict
from urllib.parse import urlparse

import openpyxl

EXCEL_FILE = os.path.join(os.path.dirname(__file__), "wiki_images.xlsx")
DEFAULT_DOCS_ROOT = os.path.join(os.path.dirname(__file__), "docs")


# ---------------------------------------------------------------------------
# Data loading
# ---------------------------------------------------------------------------

def load_data(excel_file):
    """
    Returns:
        page_images : dict  page_url -> list of (image_name, clean_image_url)
        url_to_folder: dict  clean_image_url -> dest_folder  (from Download Plan)
        url_to_skip  : dict  clean_image_url -> skip_reason or ''
    """
    wb = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)

    # Sheet 1: Wiki Images  — page_url, image_name, image_url
    ws1 = wb["Wiki Images"]
    page_images = defaultdict(list)
    seen_per_page = defaultdict(set)
    for page_url, image_name, image_url in ws1.iter_rows(min_row=2, values_only=True):
        if not page_url or not image_url:
            continue
        clean = image_url.split("?")[0]
        key = (page_url, clean)
        if key in seen_per_page:
            continue
        seen_per_page[key].add(key)
        page_images[page_url].append((image_name, clean))

    # Sheet 3: Download Plan — image_url, image_name, dest_folder, skip, skip_reason
    ws3 = wb["Download Plan"]
    url_to_folder = {}
    url_to_skip   = {}
    for row in ws3.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        clean_url   = row[0]
        dest_folder = row[2] or ""
        skip        = (row[3] == "yes")
        skip_reason = row[4] or ""
        url_to_folder[clean_url] = dest_folder
        url_to_skip[clean_url]   = skip_reason if skip else ""

    wb.close()
    return page_images, url_to_folder, url_to_skip


# ---------------------------------------------------------------------------
# RST path computation
# ---------------------------------------------------------------------------

def rst_rel_path(page_wiki_url, dest_folder, image_name, docs_root):
    """
    Compute the relative path from the migrated RST file to the image file.

    The RST file is assumed to live at:
        <docs_root>/<wiki-page-path>/index.rst

    The image lives at:
        <docs_root>/<dest_folder>/<image_name>
    """
    page_path = urlparse(page_wiki_url).path.strip("/")
    rst_dir   = os.path.join(docs_root, page_path)
    img_abs   = os.path.join(docs_root, dest_folder, image_name)
    return os.path.relpath(img_abs, rst_dir)


# ---------------------------------------------------------------------------
# Output formatting
# ---------------------------------------------------------------------------

def format_directive(rel_path, image_name, use_figure=True):
    """Return a formatted RST image or figure directive block."""
    directive = "figure" if use_figure else "image"
    lines = [
        f".. {directive}:: {rel_path}",
        f"   :align: center",
    ]
    if use_figure:
        lines += [
            "",
            f"   {image_name}",   # placeholder caption = filename
        ]
    return "\n".join(lines)


def print_page_images(page_url, images, url_to_folder, url_to_skip,
                      docs_root, use_figure=True):
    """Print all RST directives for a single wiki page."""
    page_path = urlparse(page_url).path.strip("/")
    rst_file  = f"docs/{page_path}/index.rst"

    print(f"{'=' * 72}")
    print(f"Wiki page : {page_url}")
    print(f"RST file  : {rst_file}")
    print(f"{'=' * 72}")

    if not images:
        print("  (no images found for this page)\n")
        return

    skipped   = []
    to_render = []

    for image_name, clean_url in images:
        skip_reason = url_to_skip.get(clean_url, "")
        if skip_reason:
            skipped.append((image_name, clean_url, skip_reason))
            continue
        dest_folder = url_to_folder.get(clean_url, "")
        if not dest_folder:
            skipped.append((image_name, clean_url, "no folder assigned"))
            continue
        rel_path = rst_rel_path(page_url, dest_folder, image_name, docs_root)
        to_render.append((image_name, clean_url, dest_folder, rel_path))

    if to_render:
        print(f"\nContent images ({len(to_render)}):\n")
        for image_name, clean_url, dest_folder, rel_path in to_render:
            print(format_directive(rel_path, image_name, use_figure))
            print(f"   {{'# image source': '{clean_url}'}}")
            print()

    if skipped:
        print(f"Skipped ({len(skipped)} — site chrome / external CDN):")
        for image_name, clean_url, reason in skipped:
            print(f"  [{reason}]  {image_name}  {clean_url}")
        print()


# ---------------------------------------------------------------------------
# URL resolution (exact or fuzzy)
# ---------------------------------------------------------------------------

def resolve_urls(query, page_images):
    """
    Match query against known page URLs.
    Accepts:
      - exact full URL
      - URL without scheme/host  (resources/eval/user-guides/...)
      - substring of a URL
    Returns list of matching page_urls.
    """
    # Normalise: strip trailing slash, lowercase for matching
    q = query.rstrip("/").lower()

    # Exact match first
    if query in page_images:
        return [query]

    # Try prepending the base
    full = f"https://wiki.analog.com/{q.lstrip('/')}"
    if full in page_images:
        return [full]

    # Substring match
    matches = [url for url in page_images if q in url.lower()]
    return matches


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Look up RST image directives for migrated wiki pages.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument(
        "queries",
        nargs="*",
        help="Wiki page URL(s) or substrings to look up.",
    )
    parser.add_argument(
        "--list-pages",
        nargs="?",
        const="",
        metavar="FILTER",
        help="List all pages with content images, optionally filtered by substring.",
    )
    parser.add_argument(
        "--docs-root",
        default=DEFAULT_DOCS_ROOT,
        help=f"Root of the docs/ tree (default: {DEFAULT_DOCS_ROOT})",
    )
    parser.add_argument(
        "--no-figure",
        action="store_true",
        help="Emit '.. image::' instead of '.. figure::'",
    )
    args = parser.parse_args()

    print(f"Loading {EXCEL_FILE} …", file=sys.stderr, flush=True)
    page_images, url_to_folder, url_to_skip = load_data(EXCEL_FILE)
    print(f"  {len(page_images)} pages loaded.", file=sys.stderr, flush=True)

    use_figure = not args.no_figure

    # --list-pages mode
    if args.list_pages is not None:
        filt = args.list_pages.lower()
        # Only pages that have at least one non-skipped image
        results = []
        for page_url, images in sorted(page_images.items()):
            if filt and filt not in page_url.lower():
                continue
            content_count = sum(
                1 for _, clean_url in images
                if not url_to_skip.get(clean_url, "")
                and url_to_folder.get(clean_url, "")
            )
            if content_count > 0:
                results.append((page_url, content_count))

        print(f"\n{len(results)} pages with content images"
              + (f" matching '{args.list_pages}'" if filt else "") + ":\n")
        for page_url, count in results:
            print(f"  {count:4d}  {page_url}")
        return

    # Lookup mode
    if not args.queries:
        parser.print_help()
        sys.exit(0)

    for query in args.queries:
        matched = resolve_urls(query, page_images)
        if not matched:
            print(f"\nNo pages found matching: {query!r}\n")
            continue
        if len(matched) > 1:
            print(f"\nQuery {query!r} matched {len(matched)} pages:")
            for url in matched:
                print(f"  {url}")
            print("Refine your query to select one page.\n")
            continue
        page_url = matched[0]
        images   = page_images[page_url]
        print_page_images(
            page_url, images, url_to_folder, url_to_skip,
            args.docs_root, use_figure,
        )


if __name__ == "__main__":
    main()
