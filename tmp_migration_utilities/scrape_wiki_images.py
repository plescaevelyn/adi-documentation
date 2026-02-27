#!/usr/bin/env python3
"""
Scrape wiki.analog.com for all pages and extract image names/URLs.
Outputs an Excel file: wiki_images.xlsx
"""

import re
import sys
import time
import requests
import xml.etree.ElementTree as ET
from concurrent.futures import ThreadPoolExecutor, as_completed
from urllib.parse import urljoin, urlparse
from bs4 import BeautifulSoup
import openpyxl

SITEMAP_URL = "https://wiki.analog.com/doku.php?do=sitemap"
BASE_URL = "https://wiki.analog.com"
OUTPUT_FILE = "/export/data/spop/doc_images/documentation/wiki_images.xlsx"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:142.0) "
        "Gecko/20100101 Firefox/142.0"
    ),
    "Accept-Language": "en-US,en;q=0.5",
}

MAX_WORKERS = 8       # concurrent page fetches
REQUEST_TIMEOUT = 15  # seconds per request
RETRY_LIMIT = 2


def fetch(url, session, retries=RETRY_LIMIT):
    """GET a URL, retry on transient errors. Returns (url, response_text) or (url, None)."""
    for attempt in range(retries + 1):
        try:
            r = session.get(url, timeout=REQUEST_TIMEOUT, headers=HEADERS)
            if r.status_code == 200:
                return url, r.text
            elif r.status_code in (429, 503):
                wait = 2 ** attempt
                print(f"  Rate limited ({r.status_code}) on {url}, waiting {wait}s…", flush=True)
                time.sleep(wait)
            else:
                print(f"  HTTP {r.status_code} for {url}", flush=True)
                return url, None
        except requests.RequestException as e:
            if attempt < retries:
                time.sleep(2 ** attempt)
            else:
                print(f"  Error fetching {url}: {e}", flush=True)
                return url, None
    return url, None


def get_sitemap_urls(session):
    """Fetch the sitemap (gzip-compressed XML) and return a list of page URLs."""
    import gzip
    print(f"Fetching sitemap: {SITEMAP_URL}", flush=True)
    try:
        r = session.get(SITEMAP_URL, timeout=REQUEST_TIMEOUT, headers=HEADERS)
        r.raise_for_status()
    except requests.RequestException as e:
        print(f"ERROR: Could not fetch sitemap: {e}", file=sys.stderr)
        sys.exit(1)

    # The sitemap is served as application/x-gzip — decompress raw bytes
    raw = r.content
    try:
        xml_bytes = gzip.decompress(raw)
    except Exception:
        xml_bytes = raw  # already plain if decompression fails

    xml_text = xml_bytes.decode("utf-8", errors="replace")

    try:
        root = ET.fromstring(xml_text)
    except ET.ParseError:
        # Strip namespace to simplify parsing
        xml_text_clean = re.sub(r' xmlns="[^"]+"', '', xml_text, count=1)
        root = ET.fromstring(xml_text_clean)

    ns = {'sm': 'http://www.sitemaps.org/schemas/sitemap/0.9'}
    locs = root.findall('.//sm:loc', ns)
    if not locs:
        locs = root.findall('.//loc')

    urls = [loc.text.strip() for loc in locs if loc.text and loc.text.strip()]
    print(f"Found {len(urls)} URLs in sitemap.", flush=True)
    return urls


def extract_images(page_url, html):
    """Parse HTML and return a list of image name strings found on the page."""
    soup = BeautifulSoup(html, "html.parser")
    images = []
    seen = set()

    for img in soup.find_all("img"):
        src = img.get("src") or img.get("data-src") or ""
        if not src:
            continue
        # Make absolute
        abs_src = urljoin(page_url, src)
        if abs_src in seen:
            continue
        seen.add(abs_src)

        # Extract just the filename from the path
        path = urlparse(abs_src).path
        filename = path.split("/")[-1]

        images.append({
            "image_url": abs_src,
            "image_name": filename,
        })

    # Also catch images referenced in inline CSS background-image / srcset
    for tag in soup.find_all(True):
        srcset = tag.get("srcset", "")
        if srcset:
            for part in srcset.split(","):
                part_url = part.strip().split()[0]
                abs_src = urljoin(page_url, part_url)
                if abs_src not in seen:
                    seen.add(abs_src)
                    path = urlparse(abs_src).path
                    filename = path.split("/")[-1]
                    images.append({"image_url": abs_src, "image_name": filename})

    return images


def scrape_page(args):
    """Worker: fetch a page and return (page_url, images_list)."""
    page_url, session = args
    _, html = fetch(page_url, session)
    if not html:
        return page_url, []
    images = extract_images(page_url, html)
    return page_url, images


def style_header(ws, col_widths):
    """Apply bold white-on-dark-blue styling to header row and set column widths."""
    from openpyxl.styles import Font, PatternFill, Alignment
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
    ws.freeze_panes = "A2"


def assign_folder(image_url):
    """
    Given an image URL, return the destination folder path (relative, no leading slash)
    where the image should be downloaded, or None if the image should be skipped.

    Rules:
      - /lib/ images (site chrome) → None (skip)
      - external hosts (not wiki.analog.com) → None (skip)
      - /_media/resources/eval/*       → resources/eval/<sub>/images
      - /_media/resources/fpga/*       → resources/fpga/<sub>/images
      - /_media/resources/tools-software/* → resources/tools-software/<sub>/images
      - /_media/university/*           → university/<sub>/images
      - /_media/resources/<other>/*    → resources/<other>/images
      - anything else under /_media/   → images  (root catch-all)

    A path segment that contains '.' is treated as a filename artifact and is
    ignored when determining the folder depth.
    """
    if not image_url:
        return None

    parsed = urlparse(image_url)

    # Skip non-wiki or lib/chrome images
    if parsed.netloc != "wiki.analog.com":
        return None
    if parsed.path.startswith("/lib/"):
        return None
    if "/_media/" not in parsed.path:
        return None

    # Strip query string and /_media/ prefix, then split into segments
    inner = parsed.path.replace("/_media/", "")
    # Remove filename (last segment); also skip any segment that looks like a filename
    segments = [s for s in inner.split("/") if s and "." not in s]
    # segments is now the folder path components only (no filename, no dot-artifacts)

    # Depth-3 categories: resources/eval, resources/fpga, resources/tools-software
    DEPTH3 = {"eval", "fpga", "tools-software"}

    # Normalise typo variant: 'user-guide' → 'user-guides'
    ALIASES = {"user-guide": "user-guides"}

    if not segments:
        return "images"

    top = segments[0]  # 'resources' or 'university' or other

    if top == "university":
        # depth-2: university/<sub>
        if len(segments) >= 2:
            folder = f"university/{segments[1]}/images"
        else:
            folder = "university/images"

    elif top == "resources":
        if len(segments) < 2:
            return "resources/images"
        sub = segments[1]  # eval, fpga, tools-software, etc.
        if sub in DEPTH3 and len(segments) >= 3:
            sub3 = ALIASES.get(segments[2], segments[2])
            folder = f"resources/{sub}/{sub3}/images"
        else:
            folder = f"resources/{sub}/images"

    else:
        # Any other top-level namespace under _media
        folder = f"{top}/images"

    return folder


def common_path_ancestor(pages):
    """
    Given a list of wiki page URLs, compute the common path ancestor of their
    parent directories and return it as '<common_ancestor>/images'.

    Example:
      pages = ['https://wiki.analog.com/user-guide/work/A',
               'https://wiki.analog.com/user-guide/B']
      parent dirs → ['user-guide/work', 'user-guide']
      common ancestor segments → ['user-guide']
      result → 'user-guide/images'

    If the common ancestor is the root (empty), returns 'images'.
    """
    from urllib.parse import urlparse

    # Get the parent directory of each page path (strip the page name = last segment)
    dirs = []
    for p in pages:
        path = urlparse(p).path.strip("/")
        segments = path.split("/")
        # parent = all segments except the last (the page name itself)
        parent_segments = segments[:-1] if len(segments) > 1 else []
        dirs.append(parent_segments)

    # Find common prefix segment-by-segment
    if not dirs:
        return "images"
    common = dirs[0]
    for d in dirs[1:]:
        # Zip and keep only matching leading segments
        new_common = []
        for a, b in zip(common, d):
            if a == b:
                new_common.append(a)
            else:
                break
        common = new_common

    ancestor = "/".join(common)
    return f"{ancestor}/images" if ancestor else "images"


def build_shared_images(rows):
    """
    From all_rows, find images referenced on more than one wiki page.
    Returns a list of (image_name, image_url, common_ancestor, pages_list)
    sorted by number of pages descending.
    """
    from collections import defaultdict
    # Key by image_url so the same file isn't double-counted per URL
    img_pages = defaultdict(lambda: {"image_name": "", "pages": set()})
    for page_url, image_name, image_url in rows:
        img_pages[image_url]["image_name"] = image_name
        img_pages[image_url]["pages"].add(page_url)

    # Keep only images that appear on 2+ pages
    shared = []
    for image_url, data in img_pages.items():
        if len(data["pages"]) > 1:
            pages = sorted(data["pages"])
            ancestor = common_path_ancestor(pages)
            shared.append((data["image_name"], image_url, ancestor, pages))

    # Sort by number of pages descending
    shared.sort(key=lambda x: len(x[3]), reverse=True)
    return shared


def write_excel(rows, output_file):
    """
    Write two sheets to Excel:
      Sheet 1 – 'Wiki Images': every (page_url, image_name, image_url) row.
      Sheet 2 – 'Shared Images': images referenced on 2+ pages, one row per image,
                 with page URLs spread across columns.
    rows = list of (page_url, image_name, image_url)
    """
    wb = openpyxl.Workbook()

    # --- Sheet 1: all image references ---
    ws1 = wb.active
    ws1.title = "Wiki Images"
    ws1.append(["Wiki Page URL", "Image Name", "Image URL"])
    style_header(ws1, [80, 60, 100])
    for row in rows:
        ws1.append(list(row))

    # --- Sheet 2: shared images (referenced on 2+ pages) ---
    shared = build_shared_images(rows)
    print(f"Images referenced on 2+ pages: {len(shared)}", flush=True)

    ws2 = wb.create_sheet(title="Shared Images")
    # Dynamic header: fixed columns + Wiki Page 1, Wiki Page 2, …
    max_pages = max(len(s[3]) for s in shared) if shared else 0
    header = ["Image Name", "Image URL", "Page Count", "Common Ancestor Path"] + [
        f"Wiki Page {i+1}" for i in range(max_pages)
    ]
    ws2.append(header)
    # Column widths: name=60, url=100, count=12, ancestor=60, then 80 per page column
    col_widths = [60, 100, 12, 60] + [80] * max_pages
    style_header(ws2, col_widths)

    for image_name, image_url, ancestor, pages in shared:
        ws2.append([image_name, image_url, len(pages), ancestor] + pages)

    wb.save(output_file)
    print(f"Saved: {output_file}", flush=True)


def main():
    session = requests.Session()
    session.headers.update(HEADERS)

    # 1. Get all page URLs from sitemap
    page_urls = get_sitemap_urls(session)

    if not page_urls:
        print("No URLs found in sitemap. Exiting.", file=sys.stderr)
        sys.exit(1)

    # 2. Scrape pages concurrently
    print(f"\nScraping {len(page_urls)} pages with {MAX_WORKERS} workers…", flush=True)
    all_rows = []  # (page_url, image_name, image_url)
    done = 0
    total = len(page_urls)

    args_list = [(url, session) for url in page_urls]

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = {executor.submit(scrape_page, args): args[0] for args in args_list}
        for future in as_completed(futures):
            page_url, images = future.result()
            done += 1
            if images:
                for img in images:
                    all_rows.append((page_url, img["image_name"], img["image_url"]))
            if done % 50 == 0 or done == total:
                print(f"  Progress: {done}/{total} pages processed, {len(all_rows)} images found so far", flush=True)

    print(f"\nTotal: {len(all_rows)} image references across {total} pages.", flush=True)

    # 3. Write Excel
    write_excel(all_rows, OUTPUT_FILE)


if __name__ == "__main__":
    main()
