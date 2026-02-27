#!/usr/bin/env python3
"""
scrape_wiki_downloads.py

Crawls all wiki.analog.com pages from the sitemap and collects every
downloadable file linked from each page (PDF, ZIP, EXE, DOCX, etc.).

Outputs: wiki_downloads.xlsx  with two sheets:
  - All Downloads  : one row per (page_url, link_text, file_name, file_url, host, extension)
  - Summary        : one row per unique file URL, with all referencing pages listed
"""

import gzip
import re
import sys
import time
import xml.etree.ElementTree as ET
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from urllib.parse import urljoin, urlparse

import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

SITEMAP_URL  = "https://wiki.analog.com/doku.php?do=sitemap"
OUTPUT_FILE  = "/export/data/spop/doc_images/documentation/wiki_downloads.xlsx"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:142.0) "
        "Gecko/20100101 Firefox/142.0"
    ),
    "Accept-Language": "en-US,en;q=0.5",
}

MAX_WORKERS     = 8
REQUEST_TIMEOUT = 15
RETRY_LIMIT     = 2

# All file extensions treated as downloadable (longest-first for correct matching)
DOWNLOAD_EXTS = sorted({
    ".tar.gz", ".tar.bz2", ".tar.xz",          # multi-part first
    ".pdf", ".zip", ".tar", ".gz", ".bz2", ".xz", ".tgz", ".lzma", ".rar", ".7z",
    ".exe", ".msi", ".dmg", ".deb", ".rpm",
    ".docx", ".xlsx", ".pptx", ".doc", ".xls", ".ppt",
    ".csv", ".json", ".xml",
    ".bin", ".hex", ".elf", ".img", ".iso",
    ".mat", ".m",
    ".py", ".sh", ".bat",
    ".ldf", ".uci",                              # ADI-specific project files
}, key=len, reverse=True)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def is_download_url(abs_url):
    """Return (is_download, extension) for a URL."""
    path = urlparse(abs_url).path.lower().split("?")[0]
    for ext in DOWNLOAD_EXTS:
        if path.endswith(ext):
            return True, ext
    return False, None


def fetch(url, session, retries=RETRY_LIMIT):
    """GET url with retries. Returns response text or None."""
    for attempt in range(retries + 1):
        try:
            r = session.get(url, timeout=REQUEST_TIMEOUT, headers=HEADERS)
            if r.status_code == 200:
                return r.text
            elif r.status_code in (429, 503):
                time.sleep(2 ** attempt)
            else:
                return None
        except requests.RequestException:
            if attempt < retries:
                time.sleep(2 ** attempt)
    return None


def get_sitemap_urls(session):
    """Fetch and parse the gzip-compressed sitemap. Returns list of page URLs."""
    print(f"Fetching sitemap: {SITEMAP_URL}", flush=True)
    try:
        r = session.get(SITEMAP_URL, timeout=REQUEST_TIMEOUT, headers=HEADERS)
        r.raise_for_status()
    except requests.RequestException as e:
        print(f"ERROR: {e}", file=sys.stderr)
        sys.exit(1)

    raw = r.content
    try:
        xml_bytes = gzip.decompress(raw)
    except Exception:
        xml_bytes = raw

    xml_text = xml_bytes.decode("utf-8", errors="replace")
    try:
        root = ET.fromstring(xml_text)
    except ET.ParseError:
        xml_text = re.sub(r' xmlns="[^"]+"', "", xml_text, count=1)
        root = ET.fromstring(xml_text)

    ns   = {"sm": "http://www.sitemaps.org/schemas/sitemap/0.9"}
    locs = root.findall(".//sm:loc", ns) or root.findall(".//loc")
    urls = [loc.text.strip() for loc in locs if loc.text and loc.text.strip()]
    print(f"Found {len(urls)} pages in sitemap.", flush=True)
    return urls


def extract_downloads(page_url, html):
    """
    Parse the page HTML and return a list of dicts for every downloadable link:
      page_url, link_text, file_name, file_url, host, extension
    Deduplicates by (page_url, file_url) so the same file isn't listed twice
    per page even if linked multiple times.
    """
    soup  = BeautifulSoup(html, "html.parser")
    seen  = set()
    found = []

    for a in soup.find_all("a", href=True):
        href    = a["href"].strip()
        abs_url = urljoin(page_url, href)
        # Strip query/fragment for dedup key, keep full URL for output
        clean   = abs_url.split("?")[0].split("#")[0]

        ok, ext = is_download_url(clean)
        if not ok:
            continue

        key = (page_url, clean)
        if key in seen:
            continue
        seen.add(key)

        parsed    = urlparse(abs_url)
        file_name = parsed.path.split("/")[-1]  # filename from path
        host      = parsed.netloc
        link_text = a.get_text(separator=" ", strip=True) or ""

        found.append({
            "page_url":  page_url,
            "link_text": link_text,
            "file_name": file_name,
            "file_url":  abs_url,
            "host":      host,
            "extension": ext,
        })

    return found


def scrape_page(args):
    """Worker: returns (page_url, list_of_download_dicts)."""
    page_url, session = args
    html = fetch(page_url, session)
    if not html:
        return page_url, []
    return page_url, extract_downloads(page_url, html)


# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------

def style_header(ws, col_widths):
    hfont = Font(bold=True, color="FFFFFF")
    hfill = PatternFill("solid", fgColor="1F4E79")
    halign = Alignment(horizontal="center")
    for cell in ws[1]:
        cell.font   = hfont
        cell.fill   = hfill
        cell.alignment = halign
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


def write_excel(all_rows, output_file):
    """
    all_rows: list of dicts with keys:
        page_url, link_text, file_name, file_url, host, extension
    """
    wb = openpyxl.Workbook()

    # ------------------------------------------------------------------ Sheet 1
    ws1        = wb.active
    ws1.title  = "All Downloads"
    ws1.append(["Wiki Page URL", "Link Text", "File Name", "File URL", "Host", "Extension"])
    style_header(ws1, [85, 50, 55, 110, 35, 12])

    for row in all_rows:
        ws1.append([
            row["page_url"],
            row["link_text"],
            row["file_name"],
            row["file_url"],
            row["host"],
            row["extension"],
        ])

    # ------------------------------------------------------------------ Sheet 2
    # Group by clean file URL → collect all referencing pages
    file_pages = defaultdict(lambda: {
        "file_name": "", "host": "", "extension": "", "pages": []
    })
    for row in all_rows:
        clean = row["file_url"].split("?")[0]
        entry = file_pages[clean]
        entry["file_name"]  = row["file_name"]
        entry["host"]       = row["host"]
        entry["extension"]  = row["extension"]
        if row["page_url"] not in entry["pages"]:
            entry["pages"].append(row["page_url"])

    # Sort: files referenced on most pages first
    summary = sorted(file_pages.items(), key=lambda x: -len(x[1]["pages"]))
    max_pages = max((len(v["pages"]) for _, v in summary), default=0)

    ws2       = wb.create_sheet(title="Summary")
    ws2.append(
        ["File URL", "File Name", "Host", "Extension", "Page Count"]
        + [f"Wiki Page {i+1}" for i in range(max_pages)]
    )
    style_header(ws2, [110, 55, 35, 12, 12] + [85] * max_pages)

    # Highlight rows where the same file is referenced on 2+ pages
    multi_fill = PatternFill("solid", fgColor="E2EFDA")  # light green
    for clean_url, data in summary:
        row_data = [
            clean_url,
            data["file_name"],
            data["host"],
            data["extension"],
            len(data["pages"]),
        ] + data["pages"]
        ws2.append(row_data)
        if len(data["pages"]) > 1:
            for cell in ws2[ws2.max_row]:
                cell.fill = multi_fill

    wb.save(output_file)
    print(f"\nSaved: {output_file}", flush=True)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    session = requests.Session()
    session.headers.update(HEADERS)

    # 1. Sitemap
    page_urls = get_sitemap_urls(session)

    # 2. Scrape concurrently
    print(f"\nScraping {len(page_urls)} pages ({MAX_WORKERS} workers)…", flush=True)
    all_rows = []
    done     = 0
    total    = len(page_urls)

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = {
            executor.submit(scrape_page, (url, session)): url
            for url in page_urls
        }
        for future in as_completed(futures):
            page_url, downloads = future.result()
            done += 1
            all_rows.extend(downloads)
            if done % 100 == 0 or done == total:
                print(
                    f"  {done}/{total} pages — "
                    f"{len(all_rows)} downloadable files found so far",
                    flush=True,
                )

    unique_files = len({r["file_url"].split("?")[0] for r in all_rows})
    print(f"\nTotal: {len(all_rows)} download references, {unique_files} unique files.", flush=True)

    # 3. Write Excel
    write_excel(all_rows, OUTPUT_FILE)

    # 4. Quick breakdown by extension
    from collections import Counter
    ext_counts = Counter(r["extension"] for r in all_rows)
    host_counts = Counter(r["host"] for r in all_rows)
    print("\nBreakdown by extension:")
    for ext, count in ext_counts.most_common():
        print(f"  {count:5d}  {ext}")
    print("\nBreakdown by host:")
    for host, count in host_counts.most_common(10):
        print(f"  {count:5d}  {host}")


if __name__ == "__main__":
    main()
