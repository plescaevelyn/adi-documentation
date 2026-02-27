#!/usr/bin/env python3
"""
Read the existing wiki_images.xlsx and add/update the 'Shared Images' sheet
without re-scraping the wiki.
"""

import openpyxl
from scrape_wiki_images import build_shared_images, style_header

INPUT_FILE = "/export/data/spop/doc_images/documentation/wiki_images.xlsx"

print(f"Loading {INPUT_FILE}…", flush=True)
wb = openpyxl.load_workbook(INPUT_FILE)
ws1 = wb["Wiki Images"]

# Read all rows (skip header)
rows = []
for row in ws1.iter_rows(min_row=2, values_only=True):
    if row[0]:  # skip blank rows
        rows.append(row)  # (page_url, image_name, image_url)

print(f"Loaded {len(rows)} image references.", flush=True)

# Remove old sheet if it exists
if "Shared Images" in wb.sheetnames:
    del wb["Shared Images"]

# Build shared images data
shared = build_shared_images(rows)
print(f"Images referenced on 2+ pages: {len(shared)}", flush=True)

ws2 = wb.create_sheet(title="Shared Images")
max_pages = max(len(s[2]) for s in shared) if shared else 0
header = ["Image Name", "Image URL", "Page Count"] + [
    f"Wiki Page {i+1}" for i in range(max_pages)
]
ws2.append(header)
col_widths = [60, 100, 12] + [80] * max_pages
style_header(ws2, col_widths)

for image_name, image_url, pages in shared:
    ws2.append([image_name, image_url, len(pages)] + pages)

wb.save(INPUT_FILE)
print(f"Done. Saved: {INPUT_FILE}", flush=True)
