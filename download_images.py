#!/usr/bin/env python3
"""
download_images.py

Reads wiki_images.xlsx (Sheet 1: 'Wiki Images'), assigns every unique _media
image URL to a destination folder, writes a 'Download Plan' sheet to the Excel,
then downloads all images into the docs/ folder tree.

Folder tree rooted at BASE_DIR (default: docs/ inside the repo).
Example result: docs/resources/eval/user-guides/images/foo.png

Usage:
    python3 download_images.py              # build plan + download
    python3 download_images.py --plan-only  # only write the Download Plan sheet
"""

import os
import sys
import time
import argparse
from collections import defaultdict
from urllib.parse import urlparse

import requests
import openpyxl

from scrape_wiki_images import assign_folder, style_header, HEADERS

EXCEL_FILE = "/export/data/spop/doc_images/documentation/wiki_images.xlsx"
BASE_DIR   = "/export/data/spop/doc_images/documentation/docs"

MAX_WORKERS    = 8
REQUEST_TIMEOUT = 20
RETRY_LIMIT    = 2


# ---------------------------------------------------------------------------
# Plan building
# ---------------------------------------------------------------------------

def build_download_plan(rows):
    """
    De-duplicate image URLs (strip query string), assign each a folder,
    and decide whether to skip.

    Returns a list of dicts:
      image_url      – clean URL (no query string)
      image_name     – filename
      folder         – relative destination folder, or None if skipped
      skip           – True/False
      skip_reason    – human-readable reason, or ''
    """
    # De-duplicate by clean URL
    seen = {}  # clean_url -> dict
    for page_url, image_name, image_url in rows:
        if not image_url:
            continue
        clean_url = image_url.split("?")[0]
        if clean_url in seen:
            continue

        folder = assign_folder(clean_url)
        if folder is None:
            # Determine why
            parsed = urlparse(clean_url)
            if parsed.netloc != "wiki.analog.com":
                reason = f"external host ({parsed.netloc})"
            elif parsed.path.startswith("/lib/"):
                reason = "site chrome (/lib/)"
            elif "/_media/" not in parsed.path:
                reason = "not a _media image"
            else:
                reason = "unmatched path"
            seen[clean_url] = {
                "image_url":   clean_url,
                "image_name":  image_name,
                "folder":      "",
                "skip":        True,
                "skip_reason": reason,
            }
        else:
            seen[clean_url] = {
                "image_url":   clean_url,
                "image_name":  image_name,
                "folder":      folder,
                "skip":        False,
                "skip_reason": "",
            }

    plan = list(seen.values())
    # Sort: downloadable first (by folder), then skipped
    plan.sort(key=lambda x: (x["skip"], x["folder"], x["image_name"]))
    return plan


def write_plan_sheet(wb, plan):
    """Add or replace the 'Download Plan' sheet in the workbook."""
    if "Download Plan" in wb.sheetnames:
        del wb["Download Plan"]

    ws = wb.create_sheet(title="Download Plan")
    ws.append(["Image URL", "Image Name", "Destination Folder", "Skip", "Skip Reason"])
    style_header(ws, [110, 60, 80, 8, 40])

    from openpyxl.styles import PatternFill
    skip_fill = PatternFill("solid", fgColor="FCE4D6")  # light orange for skipped rows

    for item in plan:
        row = [
            item["image_url"],
            item["image_name"],
            item["folder"],
            "yes" if item["skip"] else "no",
            item["skip_reason"],
        ]
        ws.append(row)
        if item["skip"]:
            for cell in ws[ws.max_row]:
                cell.fill = skip_fill

    return ws


# ---------------------------------------------------------------------------
# Downloading
# ---------------------------------------------------------------------------

def download_one(item, base_dir, session, retries=RETRY_LIMIT):
    """Download a single image. Returns (image_url, dest_path, status, error)."""
    url  = item["image_url"]
    name = item["image_name"]
    dest_dir  = os.path.join(base_dir, item["folder"])
    dest_path = os.path.join(dest_dir, name)

    # Skip if already downloaded
    if os.path.exists(dest_path):
        return url, dest_path, "already_exists", ""

    os.makedirs(dest_dir, exist_ok=True)

    for attempt in range(retries + 1):
        try:
            r = session.get(url, timeout=REQUEST_TIMEOUT, headers=HEADERS, stream=True)
            if r.status_code == 200:
                with open(dest_path, "wb") as f:
                    for chunk in r.iter_content(chunk_size=65536):
                        f.write(chunk)
                return url, dest_path, "ok", ""
            elif r.status_code in (429, 503):
                wait = 2 ** attempt
                time.sleep(wait)
            else:
                return url, dest_path, "error", f"HTTP {r.status_code}"
        except requests.RequestException as e:
            if attempt < retries:
                time.sleep(2 ** attempt)
            else:
                return url, dest_path, "error", str(e)

    return url, dest_path, "error", "max retries exceeded"


def download_all(plan, base_dir):
    """Download all non-skipped images concurrently. Returns summary counts."""
    from concurrent.futures import ThreadPoolExecutor, as_completed

    to_download = [item for item in plan if not item["skip"]]
    total = len(to_download)
    print(f"\nDownloading {total} images into {base_dir} …", flush=True)

    counts = {"ok": 0, "already_exists": 0, "error": 0}
    errors = []
    done = 0

    session = requests.Session()
    session.headers.update(HEADERS)

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = {
            executor.submit(download_one, item, base_dir, session): item
            for item in to_download
        }
        for future in as_completed(futures):
            url, dest_path, status, error = future.result()
            counts[status] = counts.get(status, 0) + 1
            if status == "error":
                errors.append((url, error))
            done += 1
            if done % 100 == 0 or done == total:
                print(
                    f"  {done}/{total} — ok:{counts['ok']}  "
                    f"exists:{counts['already_exists']}  "
                    f"errors:{counts['error']}",
                    flush=True,
                )

    return counts, errors


# ---------------------------------------------------------------------------
# Folder summary
# ---------------------------------------------------------------------------

def print_folder_summary(plan, base_dir):
    folder_counts = defaultdict(int)
    for item in plan:
        if not item["skip"]:
            folder_counts[item["folder"]] += 1

    print("\nImages per destination folder:")
    for folder, count in sorted(folder_counts.items(), key=lambda x: -x[1]):
        print(f"  {count:5d}  {os.path.join(base_dir, folder)}/")

    print(f"\n  Total folders : {len(folder_counts)}")
    print(f"  Total to download : {sum(folder_counts.values())}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--plan-only",
        action="store_true",
        help="Only write the Download Plan sheet; do not download.",
    )
    args = parser.parse_args()

    # 1. Load Sheet 1
    print(f"Loading {EXCEL_FILE} …", flush=True)
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws1 = wb["Wiki Images"]
    rows = [r for r in ws1.iter_rows(min_row=2, values_only=True) if r[0]]
    print(f"  Loaded {len(rows)} image references.", flush=True)

    # 2. Build plan
    plan = build_download_plan(rows)
    to_dl   = sum(1 for p in plan if not p["skip"])
    to_skip = sum(1 for p in plan if p["skip"])
    print(f"  Unique images to download : {to_dl}", flush=True)
    print(f"  Unique images to skip     : {to_skip}", flush=True)

    # 3. Write Download Plan sheet
    write_plan_sheet(wb, plan)
    wb.save(EXCEL_FILE)
    print(f"  'Download Plan' sheet written to {EXCEL_FILE}", flush=True)

    if args.plan_only:
        print("\n--plan-only: done.", flush=True)
        return

    # 4. Show folder summary before downloading
    print_folder_summary(plan, BASE_DIR)

    # 5. Download
    counts, errors = download_all(plan, BASE_DIR)

    # 6. Final report
    print("\n=== Download Report ===")
    print(f"  Downloaded  : {counts.get('ok', 0)}")
    print(f"  Already had : {counts.get('already_exists', 0)}")
    print(f"  Errors      : {counts.get('error', 0)}")

    if errors:
        print("\nFailed URLs:")
        for url, err in errors[:50]:
            print(f"  [{err}]  {url}")
        if len(errors) > 50:
            print(f"  … and {len(errors) - 50} more")

    # Count actual folders created
    folders_created = set()
    for item in plan:
        if not item["skip"]:
            folders_created.add(item["folder"])
    print(f"\n  Destination folders : {len(folders_created)}")
    print(f"  Base directory      : {BASE_DIR}")


if __name__ == "__main__":
    main()
